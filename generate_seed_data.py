#!/usr/bin/env python3
"""
REdI Data Platform — Seed Data Generator
==========================================
Reads sample data files and generates SQL INSERT statements for initial
database population. Outputs to seed_data.sql.

Usage:
    python3 generate_seed_data.py

Reads from: /mnt/user-data/uploads/ (or ./sample_data/ if running locally)
Outputs to: ./seed_data.sql
"""

import csv
import json
import hashlib
import logging
import os
import sys
from collections import OrderedDict
from datetime import datetime, date
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
UPLOAD_DIR = Path("/mnt/user-data/uploads")
OUTPUT_FILE = Path("/home/claude/migrations/010_seed_data.sql")
HASH_SALT = "redi_platform_2026"  # Application salt for patient de-identification

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("/home/claude/seed_generation.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sql_str(val):
    """Escape a value for SQL string literal. Returns NULL for empty/None."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"

def sql_int(val):
    """Convert to SQL integer literal or NULL."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NULL"
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return "NULL"

def sql_num(val):
    """Convert to SQL numeric literal or NULL."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NULL"
    try:
        return str(float(val))
    except (ValueError, TypeError):
        return "NULL"

def sql_bool(val):
    """Convert to SQL boolean."""
    if val is None or val == "":
        return "NULL"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    return "TRUE" if str(val).lower() in ("true", "1", "yes") else "FALSE"

def sql_date(val, fmt="%d/%m/%Y"):
    """Parse a date string and return SQL date literal."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NULL"
    try:
        d = datetime.strptime(val.strip(), fmt)
        return f"'{d.strftime('%Y-%m-%d')}'"
    except (ValueError, TypeError):
        return "NULL"

def hash_patient(identifier):
    """SHA-256 hash of patient identifier with salt."""
    if not identifier or str(identifier).strip() == "":
        return "NULL"
    h = hashlib.sha256(f"{HASH_SALT}:{identifier}".encode()).hexdigest()
    return f"'{h}'"

def age_to_band(age):
    """Convert numeric age to banded string."""
    if age is None:
        return "NULL"
    try:
        a = int(float(age))
        if a >= 90:
            return "'90+'"
        band_start = (a // 10) * 10
        return f"'{band_start}-{band_start + 9}'"
    except (ValueError, TypeError):
        return "NULL"

def read_csv_skipping_blanks(path):
    """Read CSV, skipping BOM and blank header lines."""
    rows = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
        # Find the first non-empty line (header)
        header_idx = 0
        for i, line in enumerate(lines):
            stripped = line.strip().strip('\r\n')
            if stripped:
                header_idx = i
                break
        reader = csv.DictReader(lines[header_idx:])
        for row in reader:
            rows.append(row)
    return rows

def map_stream(source_value):
    """Map source discipline/stream string to normalised code."""
    mapping = {
        "Medical": "medical",
        "Nursing & Midwifery": "nursing",
        "Nursing or Midwifery": "nursing",
        "Nursing": "nursing",
        "Health Practitioner": "allied_health",
        "Allied Health": "allied_health",
        "Health Practitioners": "allied_health",
        "Managerial and Clerical": "admin",
        "none": "other",
        "": "other",
    }
    return mapping.get(source_value, "other")

def map_job_family(source_label):
    """Map source job family label to normalised code."""
    mapping = {
        "Medical": "medical",
        "Visiting Medical Staff": "visiting_medical",
        "Registered / Clinical Nurse - Grades 5-6": "rn_cn_5_6",
        "Enrolled Nurses - Grades 3-4": "en_3_4",
        "Assistant In Nursing - Grades 1-2": "ain_1_2",
        "Nurse Manager - Grade 7-8": "nm_7_8",
        "Nurse Executive - Grade 9-13": "ne_9_13",
        "Health Practitioners": "health_prac",
        "Health Clinical Assistants": "health_clin_asst",
        "Managerial and Clerical": "admin_clerical",
    }
    return mapping.get(source_label)

def map_booking_status(source_value):
    """Map source booking status to normalised code."""
    mapping = {
        "Finalised": "finalised",
        "Attended": "attended",
        "Completed": "completed",
        "Enrolled": "enrolled",
        "Booked": "booked",
        "Did Not Attend": "did_not_attend",
        "Further Assessment Required": "further_assessment",
        "Cancel Request": "cancel_request",
        "Rejected": "rejected",
    }
    return mapping.get(source_value, "enrolled")

def map_course_type(source_value):
    """Map source course type to normalised code."""
    mapping = {
        "Full Course": "full_course",
        "Assessment": "assessment",
        "Refresher": "refresher",
        "ANZCA Refresher": "anzca_refresher",
        "Sim Workshop": "sim_workshop",
    }
    return mapping.get(source_value)

def map_course_status(source_value):
    """Map source course status to normalised code."""
    mapping = {
        "Open": "open",
        "Closed": "closed",
        "Cancelled": "cancelled",
    }
    return mapping.get(source_value, "closed")

def map_cert_status(source_value):
    """Map certification status to normalised code."""
    mapping = {
        "Acquired": "acquired",
        "Overdue": "overdue",
        "Assigned": "assigned",
        "Expired": "expired",
        "In Progress": "in_progress",
    }
    return mapping.get(source_value, "assigned")

# ---------------------------------------------------------------------------
# Main generation
# ---------------------------------------------------------------------------
def main():
    out_lines = []
    out = out_lines.append

    out("-- ============================================================================")
    out("-- Migration 010: Seed Data (generated from sample files)")
    out("-- REdI Data Platform")
    out(f"-- Generated: {datetime.now().isoformat()}")
    out("-- ============================================================================")
    out("")
    out("BEGIN;")
    out("")

    # ========================================================================
    # 1. ORG UNITS
    # ========================================================================
    log.info("Processing org units...")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(UPLOAD_DIR / "orgunits.xlsx")
        ws = wb.active
        ou_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                ou_rows.append(row)
        log.info(f"  Found {len(ou_rows)} org units")

        out("-- ========================================================================")
        out("-- ORG UNITS")
        out("-- ========================================================================")
        out("INSERT INTO core.org_units (id, name, directorate, service_line) VALUES")
        values = []
        for row in ou_rows:
            values.append(
                f"  ({sql_int(row[0])}, {sql_str(row[1])}, {sql_str(row[2])}, {sql_str(row[3])})"
            )
        out(",\n".join(values))
        out("ON CONFLICT (id) DO UPDATE SET")
        out("  name = EXCLUDED.name,")
        out("  directorate = EXCLUDED.directorate,")
        out("  service_line = EXCLUDED.service_line,")
        out("  updated_at = NOW();")
        out("")
    except Exception as e:
        log.error(f"  Error processing org units: {e}")

    # ========================================================================
    # 2. WARDS (extracted from inpatient census)
    # ========================================================================
    log.info("Processing wards from inpatient census...")
    try:
        wb2 = openpyxl.load_workbook(UPLOAD_DIR / "PF_Current_RBWH_Inpatients.xlsx")
        ws2 = wb2.active
        wards = set()
        ward_units_map = {}
        units = set()
        for row in ws2.iter_rows(min_row=2, values_only=True):
            ward = str(row[0]) if row[0] else None
            unit = str(row[7]) if row[7] else None
            if ward:
                wards.add(ward)
            if unit:
                units.add(unit)
            if ward and unit:
                if ward not in ward_units_map:
                    ward_units_map[ward] = set()
                ward_units_map[ward].add(unit)
        log.info(f"  Found {len(wards)} wards, {len(units)} admitting units")

        out("-- ========================================================================")
        out("-- WARDS")
        out("-- ========================================================================")
        out("INSERT INTO core.wards (code) VALUES")
        ward_values = [f"  ({sql_str(w)})" for w in sorted(wards)]
        out(",\n".join(ward_values))
        out("ON CONFLICT (code) DO NOTHING;")
        out("")

        # 3. ADMITTING UNITS
        out("-- ========================================================================")
        out("-- ADMITTING UNITS")
        out("-- ========================================================================")
        # Also pull units from transfers for division/subdivision data
        try:
            wb_tx = openpyxl.load_workbook(
                UPLOAD_DIR / "RBWH_PrevDay_Transfers_from_Other_Hospitals_to_Facility-2026-02-03.xlsx"
            )
            ws_tx = wb_tx.active
            unit_divisions = {}
            for row in ws_tx.iter_rows(min_row=6, values_only=True):
                if row[5]:  # AdmUnit
                    code = str(row[5])
                    units.add(code)
                    unit_divisions[code] = {
                        "division": str(row[6]) if row[6] else None,
                        "subdivision": str(row[7]) if row[7] else None,
                    }
        except Exception:
            unit_divisions = {}

        out("INSERT INTO core.admitting_units (code, division, subdivision) VALUES")
        unit_values = []
        for u in sorted(units):
            div_info = unit_divisions.get(u, {})
            unit_values.append(
                f"  ({sql_str(u)}, {sql_str(div_info.get('division'))}, {sql_str(div_info.get('subdivision'))})"
            )
        out(",\n".join(unit_values))
        out("ON CONFLICT (code) DO UPDATE SET")
        out("  division = COALESCE(EXCLUDED.division, core.admitting_units.division),")
        out("  subdivision = COALESCE(EXCLUDED.subdivision, core.admitting_units.subdivision);")
        out("")

        # 4. WARD-UNIT MAP
        out("-- ========================================================================")
        out("-- WARD ↔ UNIT MAP")
        out("-- ========================================================================")
        out("INSERT INTO core.ward_unit_map (ward_id, admitting_unit_id)")
        wum_values = []
        for ward_code, unit_codes in sorted(ward_units_map.items()):
            for unit_code in sorted(unit_codes):
                wum_values.append(
                    f"  ((SELECT id FROM core.wards WHERE code = {sql_str(ward_code)}), "
                    f"(SELECT id FROM core.admitting_units WHERE code = {sql_str(unit_code)}))"
                )
        out(",\n".join(wum_values))
        out("ON CONFLICT (ward_id, admitting_unit_id) DO NOTHING;")
        out("")

    except Exception as e:
        log.error(f"  Error processing wards/units: {e}")

    # ========================================================================
    # 5. STAFF (from all sources)
    # ========================================================================
    log.info("Collecting staff from all sources...")
    staff_map = OrderedDict()  # payroll_id -> {fields}

    # From ALS cert
    als_rows = read_csv_skipping_blanks(UPLOAD_DIR / "ALS_Cert.csv")
    for r in als_rows:
        pid = r.get("Person Person No.", "").strip()
        if pid:
            jf = r.get("Job Family Name", "").strip()
            staff_map.setdefault(pid, {})
            staff_map[pid].update({
                "given_name": r.get("Person Full Name", "").split()[-1] if r.get("Person Full Name") else None,
                "surname": r.get("Person Full Name", "").split()[0].rstrip(",") if r.get("Person Full Name") else None,
                "job_family_code": map_job_family(jf),
                "discipline_stream_code": map_stream(jf) if jf else staff_map[pid].get("discipline_stream_code"),
                "org_unit_id": r.get("Person Organisation Number", "").strip() or None,
                "manager_name": r.get("Manager Full Name", "").strip() or None,
            })

    # From BLS cert
    bls_rows = read_csv_skipping_blanks(
        UPLOAD_DIR / "_Grouped__Certification_Completion_Summary_003A_Organisation_003EPerson_003ECompletion_Status.csv"
    )
    for r in bls_rows:
        pid = r.get("Person Person No.", "").strip()
        if pid:
            jf = r.get("Job Family Name", "").strip()
            staff_map.setdefault(pid, {})
            if not staff_map[pid].get("job_family_code"):
                staff_map[pid]["job_family_code"] = map_job_family(jf)
            if not staff_map[pid].get("discipline_stream_code") and jf:
                staff_map[pid]["discipline_stream_code"] = map_stream(jf)
            if not staff_map[pid].get("org_unit_id"):
                staff_map[pid]["org_unit_id"] = r.get("Person Organisation Number", "").strip() or None
            if not staff_map[pid].get("manager_name"):
                staff_map[pid]["manager_name"] = r.get("Manager Full Name", "").strip() or None

    # From participants
    parts_rows = read_csv_skipping_blanks(UPLOAD_DIR / "Participants.csv")
    for r in parts_rows:
        pid = r.get("QHPayroll", "").strip()
        if pid:
            staff_map.setdefault(pid, {})
            staff_map[pid].update({
                "given_name": r.get("GivenName", "").strip() or staff_map[pid].get("given_name"),
                "surname": r.get("Surname", "").strip() or staff_map[pid].get("surname"),
                "email": r.get("Mail", "").strip() or staff_map[pid].get("email"),
                "discipline_stream_code": map_stream(r.get("Stream", "")) or staff_map[pid].get("discipline_stream_code"),
                "facility": r.get("Facility", "").strip() or staff_map[pid].get("facility"),
            })

    # From faculty
    fac_rows = read_csv_skipping_blanks(UPLOAD_DIR / "FacultyList.csv")
    for r in fac_rows:
        pid = r.get("Payroll", "").strip()
        if pid:
            staff_map.setdefault(pid, {})
            staff_map[pid].update({
                "given_name": r.get("GivenName", "").strip() or staff_map[pid].get("given_name"),
                "surname": r.get("Surname", "").strip() or staff_map[pid].get("surname"),
                "email": r.get("Mail", "").strip() or staff_map[pid].get("email"),
                "discipline_stream_code": map_stream(r.get("Stream", "")) or staff_map[pid].get("discipline_stream_code"),
            })

    log.info(f"  Total unique staff: {len(staff_map)}")

    out("-- ========================================================================")
    out("-- STAFF")
    out("-- ========================================================================")
    out("INSERT INTO core.staff (payroll_id, given_name, surname, email, discipline_stream_code, job_family_code, org_unit_id, facility, manager_name) VALUES")
    staff_values = []
    for pid, info in staff_map.items():
        org_id = info.get("org_unit_id")
        org_sql = sql_int(org_id) if org_id else "NULL"
        staff_values.append(
            f"  ({sql_str(pid)}, {sql_str(info.get('given_name'))}, {sql_str(info.get('surname'))}, "
            f"{sql_str(info.get('email'))}, {sql_str(info.get('discipline_stream_code', 'other'))}, "
            f"{sql_str(info.get('job_family_code'))}, {org_sql}, "
            f"{sql_str(info.get('facility'))}, {sql_str(info.get('manager_name'))})"
        )
    out(",\n".join(staff_values))
    out("ON CONFLICT (payroll_id) DO UPDATE SET")
    out("  given_name = COALESCE(EXCLUDED.given_name, core.staff.given_name),")
    out("  surname = COALESCE(EXCLUDED.surname, core.staff.surname),")
    out("  email = COALESCE(EXCLUDED.email, core.staff.email),")
    out("  discipline_stream_code = COALESCE(EXCLUDED.discipline_stream_code, core.staff.discipline_stream_code),")
    out("  job_family_code = COALESCE(EXCLUDED.job_family_code, core.staff.job_family_code),")
    out("  org_unit_id = COALESCE(EXCLUDED.org_unit_id, core.staff.org_unit_id),")
    out("  facility = COALESCE(EXCLUDED.facility, core.staff.facility),")
    out("  manager_name = COALESCE(EXCLUDED.manager_name, core.staff.manager_name),")
    out("  last_seen_at = NOW(),")
    out("  updated_at = NOW();")
    out("")

    # ========================================================================
    # 6. COURSES
    # ========================================================================
    log.info("Processing courses...")
    events = read_csv_skipping_blanks(UPLOAD_DIR / "Events.csv")
    out("-- ========================================================================")
    out("-- COURSES")
    out("-- ========================================================================")
    out("INSERT INTO training.courses (source_id, title, course_type_code, course_date, start_time, end_time, duration_hours, venue, capacity, status_code, outlook_id) VALUES")
    course_values = []
    for r in events:
        sid = r.get("ID", "").strip()
        if not sid:
            continue
        start = r.get("CourseStart", "").strip()
        end = r.get("CourseEnd", "").strip()
        # Calculate duration
        dur = "NULL"
        if start and end:
            try:
                s = datetime.strptime(start, "%H:%M")
                e = datetime.strptime(end, "%H:%M")
                dur = str(round((e - s).seconds / 3600, 2))
            except ValueError:
                pass
        course_values.append(
            f"  ({sql_int(sid)}, {sql_str(r.get('CourseTitle'))}, "
            f"{sql_str(map_course_type(r.get('CourseType', '')))}, "
            f"{sql_date(r.get('CourseDate', ''), '%d-%b-%Y')}, "
            f"{sql_str(start) if start else 'NULL'}, {sql_str(end) if end else 'NULL'}, "
            f"{dur}, {sql_str(r.get('CourseVenue'))}, {sql_int(r.get('CourseCap'))}, "
            f"{sql_str(map_course_status(r.get('CourseStatus', '')))}, "
            f"{sql_str(r.get('OutlookID'))})"
        )
    out(",\n".join(course_values))
    out("ON CONFLICT (source_id) DO UPDATE SET")
    out("  title = EXCLUDED.title,")
    out("  course_type_code = EXCLUDED.course_type_code,")
    out("  status_code = EXCLUDED.status_code,")
    out("  updated_at = NOW();")
    out("")

    # ========================================================================
    # 7. FACULTY MEMBERS
    # ========================================================================
    log.info("Processing faculty...")
    out("-- ========================================================================")
    out("-- FACULTY MEMBERS")
    out("-- ========================================================================")
    out("INSERT INTO training.faculty_members (source_id, given_name, surname, email, mobile, payroll_id, discipline, discipline_stream_code, certification_date, is_inactive, staff_id) VALUES")
    fac_values = []
    for i, r in enumerate(fac_rows, 1):
        pid = r.get("Payroll", "").strip()
        stream = map_stream(r.get("Stream", ""))
        inactive = r.get("Inactive", "").strip().lower() == "true"
        fac_values.append(
            f"  ({i}, {sql_str(r.get('GivenName'))}, {sql_str(r.get('Surname'))}, "
            f"{sql_str(r.get('Mail'))}, {sql_str(r.get('Mobile'))}, {sql_str(pid)}, "
            f"{sql_str(r.get('Discipline'))}, {sql_str(stream)}, "
            f"{sql_date(r.get('CertificationDate', ''), '%d/%m/%Y')}, "
            f"{sql_bool(inactive)}, "
            f"(SELECT id FROM core.staff WHERE payroll_id = {sql_str(pid)}))"
        )
    out(",\n".join(fac_values))
    out("ON CONFLICT DO NOTHING;")
    out("")

    # ========================================================================
    # 8. SAMPLE ALERT RULES
    # ========================================================================
    log.info("Generating sample alert rules...")
    out("-- ========================================================================")
    out("-- SAMPLE ALERT RULES")
    out("-- ========================================================================")
    out("""
INSERT INTO system.alert_rules (name, description, domain, metric, group_by, method, threshold_value, lookback_periods) VALUES
    ('High Escalation Rate - Ward',
     'Fires when a ward''s daily escalation count exceeds 2 standard deviations above its 30-day mean',
     'escalation', 'daily_escalation_count', 'ward', 'z_score', 2.0, 30),
    ('High Escalation Rate - Unit',
     'Fires when an admitting unit''s daily escalation count exceeds 2 standard deviations above its 30-day mean',
     'escalation', 'daily_escalation_count', 'unit', 'z_score', 2.0, 30),
    ('Falling ALS Compliance',
     'Fires when an org unit''s ALS compliance shows a negative slope over 8 consecutive weeks',
     'training', 'compliance_pct', 'org_unit', 'trend_slope', -0.5, 8),
    ('Falling BLS Compliance',
     'Fires when an org unit''s BLS compliance shows a negative slope over 8 consecutive weeks',
     'training', 'compliance_pct', 'org_unit', 'trend_slope', -0.5, 8),
    ('Critical Escalation Spike',
     'Fires when total critical escalations (code blue + MET + stroke + trauma) exceed 10 in a single day',
     'escalation', 'daily_critical_count', NULL, 'threshold', 10, 1);
""")

    out("COMMIT;")
    out("")
    out("-- ============================================================================")
    out("-- POST-SEED: Verify counts")
    out("-- ============================================================================")
    out("DO $$")
    out("BEGIN")
    out("  RAISE NOTICE 'Org units: %', (SELECT COUNT(*) FROM core.org_units);")
    out("  RAISE NOTICE 'Wards: %', (SELECT COUNT(*) FROM core.wards);")
    out("  RAISE NOTICE 'Admitting units: %', (SELECT COUNT(*) FROM core.admitting_units);")
    out("  RAISE NOTICE 'Ward-unit mappings: %', (SELECT COUNT(*) FROM core.ward_unit_map);")
    out("  RAISE NOTICE 'Staff: %', (SELECT COUNT(*) FROM core.staff);")
    out("  RAISE NOTICE 'Courses: %', (SELECT COUNT(*) FROM training.courses);")
    out("  RAISE NOTICE 'Faculty: %', (SELECT COUNT(*) FROM training.faculty_members);")
    out("  RAISE NOTICE 'Alert rules: %', (SELECT COUNT(*) FROM system.alert_rules);")
    out("END $$;")

    # Write output
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))

    log.info(f"Seed data written to {OUTPUT_FILE}")
    log.info(f"Total lines: {len(out_lines)}")

if __name__ == "__main__":
    main()
